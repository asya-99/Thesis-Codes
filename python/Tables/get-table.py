from openpyxl import load_workbook
import functools
import pandas as pd


INDUSTRIES = ["B05T09", "C17T18", "C19", "C20", "C23", "C24", "D35_E36T39"]
SUFFIXES = ["coef", "se"]

def get_push_pull_worksheets(filename):
    """
        Definition: Reads an excel file, gets the push and pull worksheets.
    """
    push_worksheets = []
    pull_worksheets = []

    wb = load_workbook(filename)
    
    for each in wb.sheetnames:
        if each.find("PUSH") != -1: push_worksheets.append(each)
        else: pull_worksheets.append(each)

    print("Push Worksheets:", push_worksheets)
    print("Pull Worksheets:", pull_worksheets)

    wb.close()

    return push_worksheets, pull_worksheets

def process_push_worksheet(filename, worksheet_name):
    """
        Definition: Create a dataframe in the desired format using a push worksheet.
    """
    lag_type = worksheet_name.split()[-1]
    #if lag_type[0] == "l":
    #    lag_type = f"t-{lag_type[-1]}"
    
    if lag_type[-1] == "t":
        lag_type = "t"

    elif lag_type[-1] == "1":
        lag_type = "t-1"

    elif lag_type[-1] == "2":
        lag_type = "t-2"

    df = pd.read_excel(filename, sheet_name=worksheet_name)
    df = df[df["Industry"].isin(INDUSTRIES)]

    df[f"OECD({lag_type})"] = df.apply( # Apply the following function for each row
        lambda row: f"{row['OECD_coef']} ({row['OECD_se']})", axis=1)

    df[f"EU({lag_type})"] = df.apply(
        lambda row: f"{row['EU_coef']} ({row['EU_se']})", axis=1)

    return df[["Industry", "Description", f"OECD({lag_type})", f"EU({lag_type})"]]

def process_pull_worksheet(filename, worksheet_name):
    lag_type = worksheet_name.split()[-1]
    #if lag_type[0] == "l":
    #    lag_type = f"t-{lag_type[-1]}"
    if lag_type[-1] == "t":
            lag_type = "t"
    
    elif lag_type[-1] == "1":
        lag_type = "t-1"

    elif lag_type[-1] == "2":
        lag_type = "t-2"

    df = pd.read_excel(filename, sheet_name=worksheet_name)
    df = df[df["Industry"].isin(INDUSTRIES)]

    df[f"NON-OECD({lag_type})"] = df.apply(
        lambda row: f"{row['All_nonOECD_coef']} ({row['All_nonOECD_se']})", axis=1)
    
    df[f"NON-OECD Rents > 1pct({lag_type})"] = df.apply(
        lambda row: f"{row['Rents_1pct_coef']} ({row['Rents_1pct_se']})", axis=1)
    
    df[f"NON-OECD Rents > 2pct({lag_type})"] = df.apply(
        lambda row: f"{row['Rents_2pct_coef']} ({row['Rents_2pct_se']})", axis=1)
    
    return df[["Industry", "Description", f"NON-OECD({lag_type})", f"NON-OECD Rents > 1pct({lag_type})", f"NON-OECD Rents > 2pct({lag_type})"]]


def process_file(in_filename, worksheets, worksheet_processor, columns, out_filename):
    """
        Definition: Merge the processed worksheets into a single table
    """
    df_list = []

    for worksheet_name in worksheets:
        if worksheet_name.find("Ratio") != -1:
            new_df = worksheet_processor(in_filename, worksheet_name)
            df_list.append(new_df)

    main_df = functools.reduce(lambda left, right: pd.merge(left, right, on=["Industry", "Description"], how="outer"), df_list)
    main_df = main_df[columns]
    main_df.to_csv(out_filename, index=False)


# SPEC 1

push_worksheets, pull_worksheets = get_push_pull_worksheets("PPML_Spec1_bylags.xlsx")

process_file("PPML_Spec1_bylags.xlsx", push_worksheets, process_push_worksheet,
             ["Industry", "OECD(t)", "OECD(t-1)", "OECD(t-2)", "EU(t)", "EU(t-1)", "EU(t-2)"],
             "PUSH_EPS_DOM_GVA_RATIO_OECD_EU.csv")

process_file("PPML_Spec1_bylags.xlsx", pull_worksheets, process_pull_worksheet,
             ["Industry", "NON-OECD(t)", "NON-OECD(t-1)", "NON-OECD(t-2)",
                          "NON-OECD Rents > 1pct(t)", "NON-OECD Rents > 1pct(t-1)", "NON-OECD Rents > 1pct(t-2)",
                          "NON-OECD Rents > 2pct(t)", "NON-OECD Rents > 2pct(t-1)", "NON-OECD Rents > 2pct(t-2)"],
             "PULL_EPI_FOR_GVA_RATIO_NON_OECD_OIL_RENTS_SPEC1.csv")


push_worksheets, pull_worksheets = get_push_pull_worksheets("PPML_Spec2_bylags.xlsx")

process_file("PPML_Spec2_bylags.xlsx", push_worksheets, process_push_worksheet,
             ["Industry", "OECD(t)", "OECD(t-1)", "OECD(t-2)", "EU(t)", "EU(t-1)", "EU(t-2)"],
             "PUSH_EPI_DOM_GVA_RATIO_OECD_EU.csv")

process_file("PPML_Spec2_bylags.xlsx", pull_worksheets, process_pull_worksheet,
             ["Industry", "NON-OECD(t)", "NON-OECD(t-1)", "NON-OECD(t-2)",
                          "NON-OECD Rents > 1pct(t)", "NON-OECD Rents > 1pct(t-1)", "NON-OECD Rents > 1pct(t-2)",
                          "NON-OECD Rents > 2pct(t)", "NON-OECD Rents > 2pct(t-1)", "NON-OECD Rents > 2pct(t-2)"],
             "PULL_EPI_FOR_GVA_RATIO_NON_OECD_OIL_RENTS_SPEC2.csv")
